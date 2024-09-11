import pandas as pd
import requests
import zipfile
import io
import tempfile
import matplotlib.pyplot as plt
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image

def download_and_extract(url, keyword):
    """Download and extract the data file."""
    try:
        response = requests.get(url)
        response.raise_for_status()
        with io.BytesIO(response.content) as zip_buffer:
            with zipfile.ZipFile(zip_buffer) as zip_ref:
                target_file = next((file for file in zip_ref.namelist() if keyword in file), None)
                if target_file:
                    zip_ref.extract(target_file)
                    return target_file
                else:
                    return None
    except Exception as e:
        st.error(f"Ein Fehler ist aufgetreten: {e}")
        return None

def process_data(file_path):
    """Process the CSV data."""
    df = pd.read_csv(file_path, sep=';', usecols=["MESS_DATUM", "TT_TU"])
    df.rename(columns={"MESS_DATUM": "Zeitstempel", "TT_TU": "Wert"}, inplace=True)
    df['Zeitstempel'] = pd.to_datetime(df['Zeitstempel'], format='%Y%m%d%H')
    df['Jahr'] = df['Zeitstempel'].dt.year
    df['Monat'] = df['Zeitstempel'].dt.month
    df['Tag'] = df['Zeitstempel'].dt.date
    df['Uhrzeit'] = df['Zeitstempel'].dt.time
    return df

def create_pivot_tables(df):
    """Create pivot tables for temperature exceedances."""
    pivot_hours = df[df['Wert'] >= 27].pivot_table(
        index='Jahr', columns='Monat', values='Wert', 
        aggfunc='count', fill_value=0
    ).reindex(columns=range(1, 13), fill_value=0)
    pivot_hours.columns = ['Jan', 'Feb', 'Mrz', 'Apr', 'Mai', 'Jun', 'Jul', 'Aug', 'Sep', 'Okt', 'Nov', 'Dez']
    
    df_days = df[df['Wert'] >= 27].groupby('Tag').size().reset_index(name='Überschreitungen')
    df_days['Jahr'] = pd.to_datetime(df_days['Tag']).dt.year
    df_days['Monat'] = pd.to_datetime(df_days['Tag']).dt.month
    
    pivot_days = df_days.pivot_table(
        index='Jahr', columns='Monat', values='Überschreitungen', 
        aggfunc='count', fill_value=0
    ).reindex(columns=range(1, 13), fill_value=0)
    pivot_days.columns = ['Jan', 'Feb', 'Mrz', 'Apr', 'Mai', 'Jun', 'Jul', 'Aug', 'Sep', 'Okt', 'Nov', 'Dez']

    return pivot_hours, pivot_days

def save_monthly_data(df, writer):
    """Save individual sheets for each year and month with date, time, and temperature."""
    months = {
        1: "Jan", 2: "Feb", 3: "Mrz", 4: "Apr", 5: "Mai", 6: "Jun",
        7: "Jul", 8: "Aug", 9: "Sep", 10: "Okt", 11: "Nov", 12: "Dez"
    }

    fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')

    for year in df['Jahr'].unique():
        for month_num, month_name in months.items():
            month_data = df[(df['Jahr'] == year) & (df['Monat'] == month_num)][['Tag', 'Uhrzeit', 'Wert']]
            if not month_data.empty:
                sheet_name = f"{year}_{month_name}"
                month_data.rename(columns={'Tag': 'Datum', 'Wert': 'Temperatur'}, inplace=True)
                
                month_data['Datum'] = pd.to_datetime(month_data['Datum']).dt.strftime('%d.%m.%Y')
                
                month_data.to_excel(writer, sheet_name=sheet_name, index=False)

                workbook = writer.book
                worksheet = workbook[sheet_name]
                for col in range(len(month_data.columns)):
                    max_length = max(month_data.iloc[:, col].astype(str).map(len).max(), len(month_data.columns[col])) + 2
                    worksheet.column_dimensions[get_column_letter(col + 1)].width = max_length
                
                for row in worksheet.iter_rows(min_row=2, min_col=3, max_col=3, max_row=worksheet.max_row):
                    for cell in row:
                        if cell.value and cell.value >= 27:
                            cell.fill = fill

def plot_and_save_pivot_tables(pivot_hours, pivot_days):
    """Plot and save the pivot tables as images."""
    def plot_pivot(pivot_table, title):
        plt.figure(figsize=(12, 8))
        pivot_table.T.plot(kind='line', marker='o')  # Transponieren und Linienplot
        plt.title(title)
        plt.xlabel('Monat')
        plt.ylabel('Anzahl der Überschreitungen')
        plt.xticks(range(12), ['Jan', 'Feb', 'Mrz', 'Apr', 'Mai', 'Jun', 'Jul', 'Aug', 'Sep', 'Okt', 'Nov', 'Dez'], rotation=45)
        plt.legend(title='Jahr')
        plt.tight_layout()
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
        plt.savefig(temp_file.name, bbox_inches='tight')
        plt.close()
        return temp_file.name

    image_path_hours = plot_pivot(pivot_hours, 'Anzahl der Stunden mit Temperaturen ≥ 27°C pro Jahr und Monat')
    image_path_days = plot_pivot(pivot_days, 'Anzahl der Tage mit Temperaturen ≥ 27°C pro Jahr und Monat')

    return image_path_hours, image_path_days

def main():
    st.title('Wetterdaten Analyse')

    url = 'https://opendata.dwd.de/climate_environment/CDC/observations_germany/climate/hourly/air_temperature/recent/stundenwerte_TU_02014_akt.zip'
    keyword = 'produkt_tu_stunde'

    if st.button('Daten herunterladen und analysieren'):
        target_file = download_and_extract(url, keyword)
        if not target_file:
            st.error("Die Zieldatei konnte nicht heruntergeladen oder extrahiert werden.")
            return

        df = process_data(target_file)
        pivot_hours, pivot_days = create_pivot_tables(df)

        excel_path = 'wetterdaten_analyse_jahr_monat_linie.xlsx'
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            pivot_hours.to_excel(writer, sheet_name='Überschreitungen (Stunden)')
            pivot_days.to_excel(writer, sheet_name='Überschreitungen (Tage)')
            save_monthly_data(df, writer)
            image_path_hours, image_path_days = plot_and_save_pivot_tables(pivot_hours, pivot_days)
            workbook = writer.book
            worksheet_hours = workbook['Überschreitungen (Stunden)']
            worksheet_days = workbook['Überschreitungen (Tage)']
            img_hours = Image(image_path_hours)
            img_days = Image(image_path_days)
            worksheet_hours.add_image(img_hours, 'E5')
            worksheet_days.add_image(img_days, 'E5')

        st.success(f"Excel-Datei wurde erstellt: {excel_path}")
        with open(excel_path, 'rb') as file:
            st.download_button(label='Download Excel-Datei', data=file, file_name=excel_path)

if __name__ == "__main__":
    main()
